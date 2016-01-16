from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.corpus import wordnet

from openpyxl import load_workbook
from openpyxl import Workbook

import string

import re
filter_name = re.compile('[.][a-z][.][0-9]+', re.I)

'''import pexpect

try:
    fobj = open('last_access.txt', 'r+b')
except IOError:
    pexpect.run('touch last_access.txt')
    fobj = open('last_access.txt', 'r+b')
    data = fobj.read()
    #print data,type(data)
    if data == '':
        #print 1
        i = '2'
        fobj.write(str(int(i) + 1))
    else:
        i = str(int(data) + 1)
        fobj.write(i)
        fobj.close()
else:
    data = fobj.read()
    if data == '':
        i = '2'
        fobj.write(i)
    else:
        i = str(int(data) + 1)

        fobj.write(i)
        fobj.close()
        '''
garbage = stopwords.words('english') + list(string.punctuation)

# below code for accessing the drive and getting atuthentication and permission grant

gauth = GoogleAuth()
drive = GoogleDrive(gauth)
# for downloading the file
response = drive.CreateFile({'id': "19zNYJ0dElbYmeC7ohupht3YjhlsAdmvcGLwkirNmYOI"})
response.GetContentFile('Project Portal (Responses).xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # Save Drive file as a local file

# student response sheet and faculty database
student = load_workbook(filename = "Project Portal (Responses).xlsx")
faculty = load_workbook(filename= 'Faculty_Database.xlsx')
std_sheet = student['Form Responses 1']
fac_sheet = faculty['Sheet1']

j = '2'

#try :
 #   faculty_file = open('faculty_database.txt','r')
#except IOError:

faculty_profile = {}
while True:
    try :
        words = fac_sheet['C'+ j].value + " "+ fac_sheet['D'+ j].value + " " + fac_sheet['E' + j].value
        faculty_profile[fac_sheet['B'+ j].value] = set([i for i  in word_tokenize(words) if i not in garbage])
    except TypeError:
        try:
            words = fac_sheet['C'+ j].value + " " + fac_sheet['D'+ j].value
            faculty_profile[fac_sheet['B'+ j].value] = set([i for i  in word_tokenize(words) if i not in garbage])
        except TypeError:
            try:
                words = fac_sheet['C'+ j].value
                faculty_profile[fac_sheet['B'+ j].value] = set([i for i  in word_tokenize(words) if i not in garbage])
            except TypeError:
                break
    print words
    #faculty_profile[fac_sheet['B'+ j].value] = set([i for i  in word_tokenize(words) if i not in garbage])
    j = str(int(j) + 1)
#else:
#    data = faculty_file.read()
print(faculty_profile)


i = '5'

# faculty refer sheet creation
wb = Workbook()
ws = wb.active
ws.title = 'MatchSheet'

ws['A1'].value = 'Faculty Name'
ws['B1'].value = 'Division'
ws['C1'].value = 'Area of Interest'
ws['D1'].value = 'List of Projects'
ws['E1'].value = 'Match %'

d ='2'
cell_range = fac_sheet['B2': 'B81']



if std_sheet['B'+i].value == 'Yes':
    p_title = std_sheet['C'+i].value
    title_words = [k for k in word_tokenize(p_title) if i not in garbage]
    print(title_words)
    abstract = std_sheet['D'+ i].value
    abstract_words = [k for k in word_tokenize(abstract) if i not in garbage]
    print(abstract_words)
    try:
        if std_sheet['F' + i] == 'Yes':
            pass
    except TypeError:
        pass

else:
    count = 0
    refer_faculty = []
    topics = std_sheet['G'+i].value
    topics_words = word_tokenize(topics)
    for a in topics_words:
        w = wordnet.synsets(a)
        for c in w:
            for b in faculty_profile:

                #print c
                if str(c.name())[:-5] in faculty_profile[b]:
                    refer_faculty.append(b)
                    ws['A' + d] = b


    print (set(refer_faculty))
    print topics
wb.save('MatchSheet.xlsx')